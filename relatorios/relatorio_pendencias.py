from PIL import Image, ImageDraw, ImageFont
import textwrap
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell 
import io

# =========================
# RELATÓRIO
# =========================

from PIL import Image, ImageDraw, ImageFont
import textwrap

from PIL import Image, ImageDraw, ImageFont
import textwrap

def medir_texto(draw, texto, font):
    bbox = draw.textbbox((0, 0), texto, font=font)
    return bbox[2] - bbox[0]

def gerar_relatorio_imagem(df_f, filtro_consultor=None, filtro_loja=None,
                           filtro_data_inicial=None, filtro_data_final=None):
    cor_fundo = "white"
    cor_titulo = "black"
    cor_cabecalho = (237, 125, 49)  # laranja Excel

    # Fonte menor
    try:
        font = ImageFont.truetype("arial.ttf", 12)
        font_bold = ImageFont.truetype("arialbd.ttf", 14)
    except:
        font = ImageFont.load_default()
        font_bold = ImageFont.load_default()

    # Consultores
    if not filtro_consultor or filtro_consultor == "Todos":
        consultores = df_f["consultor"].dropna().unique()
    else:
        consultores = [filtro_consultor]

    # Colunas do relatório (usar nomes do DataFrame)
    colunas = ["Data", "Controle DAV", "Produto", "Erro", "Tipo", "Valor", "observacao"]

    # Definir larguras mínimas por coluna para não sobrepor texto
    largura_minima = {
        "Data": 80,
        "Controle DAV": 100,
        "Produto": 120,
        "Erro": 100,
        "Tipo": 80,
        "Valor": 80,
        "observacao": 200
    }

    # Calcular larguras reais
    img_tmp = Image.new("RGB", (1,1))
    draw_tmp = ImageDraw.Draw(img_tmp)
    col_widths = {}
    for col in colunas:
        col_key = col
        textos = [str(v) for v in df_f[col_key].dropna()] if col_key in df_f.columns else []
        textos.append(col if col != "observacao" else "Observação")
        max_w = max(medir_texto(draw_tmp, t, font) for t in textos)
        col_widths[col] = max(max_w + 10, largura_minima.get(col, 100))  # largura mínima

    largura = sum(col_widths.values()) + 20

    # Calcular altura total
    y = 10 + 30 + 40  # título + loja/período
    for cons in consultores:
        y += 25  # título consultor
        y += 25  # cabeçalho
        df_cons = df_f[df_f["consultor"] == cons]
        for _, row in df_cons.iterrows():
            obs_text = str(row['observacao'])
            wrapped = textwrap.wrap(obs_text, width=int(col_widths["observacao"]/8))
            y += 15 * len(wrapped) + 20  # linha + espaçamento
        y += 30  # divisória
    y += 30  # total loja
    altura = y + 20

    # Criar imagem
    img = Image.new("RGB", (largura, altura), cor_fundo)
    draw = ImageDraw.Draw(img)
    y = 10

    # Título
    draw.text((10, y), "RELATÓRIO AUDITORIA VX CASE - ES", font=font_bold, fill=cor_titulo)
    y += 30

    # Loja e período
    periodo = ""
    if filtro_data_inicial and filtro_data_final:
        periodo = f"Período: {filtro_data_inicial.strftime('%d/%m/%Y')} a {filtro_data_final.strftime('%d/%m/%Y')}"
    draw.text((10, y), f"Loja: {filtro_loja if filtro_loja else 'Não informado'}   {periodo}",
              font=font, fill=cor_titulo)
    y += 40

    # Total da loja
    total_loja = df_f['valor'].sum()

    # Para cada consultor
    for cons in consultores:
        df_cons = df_f[df_f["consultor"] == cons]

        # Título consultor
        draw.text((10, y), f"Consultor: {cons}", font=font_bold, fill=cor_titulo)
        y += 25

        # Cabeçalho
        x = 10
        for col in colunas:
            w = col_widths[col]
            draw.rectangle([x, y, x+w, y+25], fill=cor_cabecalho)
            header_name = "Observação" if col=="observacao" else col
            draw.text((x+3, y+4), header_name, font=font_bold, fill="black")
            x += w
        y += 25

        # Linhas do consultor
        for _, row in df_cons.iterrows():
            x = 10
            draw.text((x, y), str(row['data']), font=font, fill="black"); x += col_widths["Data"]
            draw.text((x, y), str(row['controle_dav']), font=font, fill="black"); x += col_widths["Controle DAV"]
            draw.text((x, y), str(row['produto']), font=font, fill="black"); x += col_widths["Produto"]
            draw.text((x, y), str(row['erro']), font=font, fill="black"); x += col_widths["Erro"]
            draw.text((x, y), str(row['tipo']), font=font, fill="black"); x += col_widths["Tipo"]

            # Valor alinhado à direita
            valor_text = f"R$ {row['valor']:.2f}"
            tw = medir_texto(draw, valor_text, font)
            draw.text((x + col_widths["Valor"] - tw - 3, y), valor_text, font=font, fill="black")
            x += col_widths["Valor"]

            # Observação com quebra de linha
            obs_text = str(row['observacao'])
            wrapped = textwrap.wrap(obs_text, width=int(col_widths["observacao"]/8))
            for i, line in enumerate(wrapped):
                draw.text((x+3, y + i*15), line, font=font, fill="black")
            y += 15 * len(wrapped)
            y += 10  # espaçamento entre linhas

    # Total da loja
    draw.text((10, y), f"Total da Loja: R$ {total_loja:.2f}", font=font_bold, fill="black")

    return img

# =========================
# RELATÓRIO EXCEL
# =========================

def gerar_relatorio_excel(df, filtro_data_inicial=None, filtro_data_final=None):
    # =========================
    # Filtrar pelo período
    # =========================
    df['data'] = pd.to_datetime(df['data'])
    if filtro_data_inicial and filtro_data_final:
        df = df[(df['data'].dt.date >= filtro_data_inicial) & 
                (df['data'].dt.date <= filtro_data_final)]

    # =========================
    # Agrupar por Loja e Consultor
    # =========================
    relatorio = df.groupby(['loja', 'consultor'], as_index=False)['valor'].sum()
    relatorio.rename(columns={'valor':'Valor Total'}, inplace=True)

    # =========================
    # Calcular 20% Valor e Total Desconto
    # =========================
    relatorio['20% Valor'] = relatorio['Valor Total'] * 0.2
    relatorio['Total Desconto'] = relatorio['20% Valor'].apply(lambda x: 50 if x > 50 else x)

    # =========================
    # Criar Excel com título
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Desconto Auditoria"

    # Título no topo
    ws.merge_cells('A1:E1')
    ws['A1'] = "Desconto Auditoria"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Adicionar Data do Período abaixo do título
    if filtro_data_inicial and filtro_data_final:
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Período: {filtro_data_inicial.strftime('%d/%m/%Y')} a {filtro_data_final.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')

    # Adicionar dados
    for r_idx, row in enumerate(dataframe_to_rows(relatorio, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 4:  # Cabeçalho
                cell.font = Font(bold=True)

    # Ajustar largura das colunas (ignorar células mescladas)
    for i, column_cells in enumerate(ws.columns, start=1):
        cells = [cell for cell in column_cells if not isinstance(cell, MergedCell)]
        if not cells:
            continue
        length = max(len(str(cell.value)) for cell in cells if cell.value is not None)
        ws.column_dimensions[get_column_letter(i)].width = length + 5

    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



